"""
BNA Contour Explorer v4
────────────────────────
Improvements over v3:
  • Vectorized IDW (10-100x faster for batch queries)
  • Click-to-query on interactive map (Plotly selection events)
  • Slope & Gradient Map (new task)
  • Basemap tile overlay (OpenStreetMap / Carto)
  • Temporal animation (animate across Z levels as time steps)
  • Excel export (openpyxl, multi-sheet)
  • Outlier / data-quality flagging (IQR + spatial gap detection)
"""

import io
import streamlit as st
import pandas as pd
import numpy as np
from scipy.spatial import KDTree
from scipy.interpolate import griddata
from scipy import stats as scipy_stats
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# ─────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="BNA Contour Explorer",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;600;700&family=JetBrains+Mono:wght@400;600&display=swap');

html, body, [class*="css"] { font-family: 'Space Grotesk', sans-serif; }
h1,h2,h3 { font-family: 'Space Grotesk', sans-serif !important; font-weight: 700; }
code, .mono { font-family: 'JetBrains Mono', monospace; }

[data-testid="metric-container"] {
    background: linear-gradient(135deg,#0f172a,#1e293b);
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 14px 18px;
    color: #f8fafc;
}
[data-testid="metric-container"] label { color: #94a3b8 !important; font-size: 11px; text-transform: uppercase; letter-spacing: .08em; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #38bdf8 !important; font-family: 'JetBrains Mono', monospace; font-size: 1.4rem; }

.task-card {
    background: linear-gradient(135deg,#0f172a 0%,#1e293b 100%);
    border: 1px solid #334155;
    border-radius: 16px;
    padding: 22px 18px;
    text-align: center;
    cursor: pointer;
    transition: all .25s;
    min-height: 130px;
    display: flex; flex-direction: column; align-items: center; justify-content: center;
}
.task-card:hover { border-color: #38bdf8; transform: translateY(-3px); box-shadow: 0 8px 32px rgba(56,189,248,.15); }
.task-card .icon { font-size: 2rem; margin-bottom: 8px; }
.task-card .title { font-weight: 700; color: #f1f5f9; font-size: .95rem; }
.task-card .desc  { color: #64748b; font-size: .75rem; margin-top: 4px; }

.section-header {
    background: linear-gradient(90deg,#0ea5e9,#6366f1);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    font-size: 1.6rem; font-weight: 700; margin-bottom: 4px;
}
.breadcrumb { color:#64748b; font-size:.85rem; margin-bottom:20px; }

.badge-warn { display:inline-block; background:#422006; border:1px solid #92400e; color:#fbbf24;
              border-radius:999px; padding:2px 10px; font-size:11px; margin:2px; }
.badge-ok   { display:inline-block; background:#052e16; border:1px solid #166534; color:#4ade80;
              border-radius:999px; padding:2px 10px; font-size:11px; margin:2px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Session state
# ─────────────────────────────────────────────────────────────
for key, val in {
    "task": None,
    "query_history": [],
    "df": None,
    "df2": None,
    "tree": None,
    "click_x": None,
    "click_y": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

COLORSCALES = ["Viridis", "Plasma", "Cividis", "RdBu", "Turbo", "Inferno", "Magma"]

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Parsing BNA …")
def parse_bna(file_bytes: bytes, filename: str) -> pd.DataFrame:
    points, errors = [], []
    lines = file_bytes.decode("utf-8", errors="replace").splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1; continue
        if line.startswith('"'):
            parts = line.split(',')
            try:
                z = float(parts[1].replace('"', '').strip())
                n = int(parts[2].strip())
                i += 1
                for _ in range(n):
                    if i < len(lines):
                        try:
                            cx, cy = lines[i].split(',')[:2]
                            points.append({'x': float(cx), 'y': float(cy), 'z': z})
                        except Exception as e:
                            errors.append(str(e))
                        i += 1
            except Exception as e:
                errors.append(str(e)); i += 1
        else:
            i += 1
    if errors:
        st.warning(f"Skipped {len(errors)} malformed line(s).")
    return pd.DataFrame(points)


# ── Vectorized IDW (replaces per-point Python loop) ──────────
def idw_single(df_vals, tree, qx, qy, k=6, p=2):
    """Single-point IDW — used for click queries."""
    d, idx = tree.query([qx, qy], k=min(k, len(df_vals)))
    d = np.asarray(d, dtype=float)
    v = df_vals[idx]
    if d[0] == 0:
        return float(v[0])
    w = 1.0 / (d ** p)
    return float(np.dot(w, v) / w.sum())


def idw_batch(df, tree, qxy: np.ndarray, k=6, p=2) -> np.ndarray:
    """
    Vectorized IDW for N query points at once.
    qxy : (N, 2) array of [x, y]
    Returns (N,) array of interpolated Z values.
    """
    k = min(k, len(df))
    dists, idxs = tree.query(qxy, k=k)          # (N, k)
    dists = dists.astype(float)
    z_vals = df['z'].values[idxs]                # (N, k)

    # Handle exact hits (distance == 0)
    exact = dists[:, 0] == 0
    weights = np.where(dists == 0, 0.0, 1.0 / (dists ** p))  # (N, k)
    # For exact hits, put all weight on the nearest
    weights[exact] = 0.0
    weights[exact, 0] = 1.0

    wsum = weights.sum(axis=1, keepdims=True)
    result = (weights * z_vals).sum(axis=1) / wsum.squeeze()
    return result


def make_heatmap(df, colorscale="Viridis", title="Contour Surface", res=300):
    sample = df if len(df) <= 60_000 else df.sample(60_000, random_state=0)
    xi = np.linspace(sample['x'].min(), sample['x'].max(), res)
    yi = np.linspace(sample['y'].min(), sample['y'].max(), res)
    XI, YI = np.meshgrid(xi, yi)
    ZI = griddata((sample['x'], sample['y']), sample['z'], (XI, YI), method='linear')
    fig = go.Figure(go.Heatmap(
        z=ZI, x=xi, y=yi,
        colorscale=colorscale,
        colorbar=dict(title="Z", thickness=14),
    ))
    fig.update_layout(title=title, xaxis_title="X", yaxis_title="Y",
                      height=520, margin=dict(l=20,r=20,t=40,b=20),
                      paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                      font=dict(color='#94a3b8'))
    return fig


def stat_summary(df):
    z = df['z']
    return {
        "Count":    len(z),
        "Min":      z.min(),
        "Max":      z.max(),
        "Mean":     z.mean(),
        "Median":   z.median(),
        "Std":      z.std(),
        "Skewness": float(scipy_stats.skew(z)),
        "Kurtosis": float(scipy_stats.kurtosis(z)),
    }


def is_latlon(df):
    """Heuristic: looks like geographic lat/lon data."""
    return (df['x'].between(-180, 180).all() and df['y'].between(-90, 90).all())


# ─────────────────────────────────────────────────────────────
# UPLOAD SCREEN
# ─────────────────────────────────────────────────────────────
def upload_screen():
    st.markdown('<h1 style="font-size:2.4rem;font-weight:800;color:#f1f5f9;margin-bottom:4px">🗺️ BNA Contour Explorer</h1>', unsafe_allow_html=True)
    st.markdown('<p style="color:#64748b;font-size:1rem;margin-bottom:32px">Upload a BNA file to unlock all analysis tools</p>', unsafe_allow_html=True)

    col_up, col_info = st.columns([2,1], gap="large")

    with col_up:
        uploaded  = st.file_uploader("Primary BNA file", type="bna", label_visibility="collapsed")
        st.caption("Drag & drop or click to browse  ·  .bna format")
        uploaded2 = st.file_uploader("Optional: second BNA file for comparison", type="bna")

    with col_info:
        st.markdown("""
        **What is a BNA file?**  
        A BNA (Atlas Boundary) file stores geographic contour/boundary data as named polylines with associated numeric values.

        **v4 new features:**
        - ⚡ Vectorized IDW (10-100× faster batch queries)
        - 🖱️ Click-to-query on interactive map
        - 📐 Slope & Gradient Map
        - 🌍 Basemap tile overlay
        - 🎞️ Temporal animation across Z levels
        - 📊 Excel export (multi-sheet)
        - 🔍 Outlier & data quality flagging
        """)

    if uploaded:
        with st.spinner("Parsing…"):
            df = parse_bna(uploaded.getvalue(), uploaded.name)
            df["source"] = uploaded.name
        if df.empty:
            st.error("No valid data found."); return
        st.session_state.df   = df
        st.session_state.tree = KDTree(df[['x','y']].values)
        if uploaded2:
            df2 = parse_bna(uploaded2.getvalue(), uploaded2.name)
            df2["source"] = uploaded2.name
            st.session_state.df2 = df2
        st.success(f"✅ Loaded **{len(df):,}** points from `{uploaded.name}`")
        st.rerun()


# ─────────────────────────────────────────────────────────────
# TASK HUB
# ─────────────────────────────────────────────────────────────
TASKS = [
    ("🗺️", "Interactive Map",      "Heatmap, point cloud, contour lines"),
    ("🎯", "Point Query",           "Click map or type coords + IDW"),
    ("📂", "Batch CSV Query",        "Query thousands of points fast"),
    ("🔵", "Radius / Zone Search",  "All points within a distance"),
    ("📏", "Transect Profile",      "Cross-section Z along a line"),
    ("🎨", "Contour Band Analysis", "Stats & area per contour level"),
    ("🔥", "Hotspot Detection",     "Find peaks, valleys & anomalies"),
    ("⚖️", "File Comparison",       "Diff two BNA files side-by-side"),
    ("🧊", "3-D Surface",           "Interactive 3-D perspective view"),
    ("📐", "Slope & Gradient",      "Surface slope, aspect & gradient"),
    ("🎞️", "Temporal Animation",   "Animate Z levels as time steps"),
    ("🔍", "Data Quality",          "Outlier & gap detection report"),
    ("📤", "Export & Report",       "CSV, Excel & stats report"),
]

def task_hub():
    df = st.session_state.df
    c_a,c_b,c_c,c_d,c_e = st.columns(5)
    for col, (label, val) in zip(
        [c_a,c_b,c_c,c_d,c_e],
        [("Points",f"{len(df):,}"),("Z levels",f"{df['z'].nunique()}"),
         ("Z min",f"{df['z'].min():.3f}"),("Z max",f"{df['z'].max():.3f}"),
         ("Std",f"{df['z'].std():.3f}")]):
        col.metric(label, val)

    st.markdown("---")
    st.markdown('<p style="color:#94a3b8;font-size:.9rem;margin-bottom:16px">Select a task to begin ↓</p>', unsafe_allow_html=True)

    rows = [TASKS[:5], TASKS[5:10], TASKS[10:]]
    for row in rows:
        cols = st.columns(5)
        for col, (icon, title, desc) in zip(cols, row):
            with col:
                st.markdown(f"""
                <div class="task-card">
                  <div class="icon">{icon}</div>
                  <div class="title">{title}</div>
                  <div class="desc">{desc}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Open", key=f"task_{title}"):
                    st.session_state.task = title
                    st.rerun()


def back_button():
    if st.button("← Back to task hub"):
        st.session_state.task = None
        st.rerun()


# ─────────────────────────────────────────────────────────────
# TASK 1 — Interactive Map  (with basemap toggle)
# ─────────────────────────────────────────────────────────────
def task_map():
    df = st.session_state.df
    st.markdown('<div class="section-header">🗺️ Interactive Map</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Interactive Map</div>', unsafe_allow_html=True)

    geo = is_latlon(df)

    c1, c2, c3, c4 = st.columns(4)
    view   = c1.selectbox("View type", ["Heatmap (interpolated)", "Point Cloud", "Contour Lines"])
    cscale = c2.selectbox("Colour scale", COLORSCALES)
    res    = c3.slider("Grid resolution", 100, 500, 250, 50)
    if geo:
        use_basemap = c4.checkbox("🌍 Basemap tiles", value=True,
                                  help="Overlay OpenStreetMap tiles (works when X/Y look like lon/lat)")
    else:
        use_basemap = False
        c4.caption("Basemap unavailable\n(data not in lon/lat range)")

    sample = df if len(df) <= 60_000 else df.sample(60_000, random_state=0)

    if use_basemap and geo:
        # Use Scattermapbox for geographic data
        basemap_style = "open-street-map"
        fig = go.Figure(go.Densitymapbox(
            lat=sample['y'], lon=sample['x'], z=sample['z'],
            radius=12,
            colorscale=cscale,
            colorbar=dict(title="Z"),
        ))
        fig.update_layout(
            mapbox=dict(style=basemap_style,
                        center=dict(lat=sample['y'].mean(), lon=sample['x'].mean()),
                        zoom=7),
            height=580, margin=dict(l=0,r=0,t=0,b=0),
            paper_bgcolor='#0f172a', font=dict(color='#94a3b8'),
        )
    elif view == "Heatmap (interpolated)":
        fig = make_heatmap(sample, cscale, res=res)
    elif view == "Point Cloud":
        fig = px.scatter(sample, x='x', y='y', color='z',
                         color_continuous_scale=cscale, opacity=.6,
                         title="Point Cloud", height=540)
        fig.update_traces(marker=dict(size=3))
        fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
    else:
        xi = np.linspace(sample['x'].min(), sample['x'].max(), res)
        yi = np.linspace(sample['y'].min(), sample['y'].max(), res)
        XI, YI = np.meshgrid(xi, yi)
        ZI = griddata((sample['x'], sample['y']), sample['z'], (XI, YI), method='linear')
        fig = go.Figure(go.Contour(
            z=ZI, x=xi, y=yi, colorscale=cscale,
            contours=dict(showlabels=True, labelfont=dict(size=10, color='white')),
            colorbar=dict(title="Z"),
        ))
        fig.update_layout(title="Contour Lines", height=540,
                          paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))

    st.plotly_chart(fig, use_container_width=True)
    if len(df) > 60_000:
        st.caption(f"Displaying 60,000-point sample of {len(df):,} total.")


# ─────────────────────────────────────────────────────────────
# TASK 2 — Point Query  (click-to-query + typed coords)
# ─────────────────────────────────────────────────────────────
def task_point_query():
    df   = st.session_state.df
    tree = st.session_state.tree
    st.markdown('<div class="section-header">🎯 Point Query</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Point Query</div>', unsafe_allow_html=True)

    st.info("**Click-to-query:** Click any point on the map below, then press **▶ Query selected point**. "
            "Or type coordinates manually.")

    # ── Mini clickable map ────────────────────────────────────
    sample = df.sample(min(10_000, len(df)), random_state=1)
    click_fig = px.scatter(
        sample, x='x', y='y', color='z',
        color_continuous_scale='Viridis', opacity=.5, height=360,
        title="Click a location to query",
    )
    click_fig.update_traces(marker=dict(size=4))
    click_fig.update_layout(
        paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
        font=dict(color='#94a3b8'),
        dragmode='select',
    )

    event = st.plotly_chart(click_fig, use_container_width=True, on_select="rerun", key="click_map")

    # Extract clicked coordinates from selection event
    clicked_x, clicked_y = None, None
    if event and event.get("selection") and event["selection"].get("points"):
        pt = event["selection"]["points"][0]
        clicked_x = pt.get("x")
        clicked_y = pt.get("y")
        st.session_state.click_x = clicked_x
        st.session_state.click_y = clicked_y

    # Coordinate inputs — pre-filled from click if available
    c1, c2, c3, c4 = st.columns(4)
    default_x = st.session_state.click_x if st.session_state.click_x is not None else float(df['x'].mean())
    default_y = st.session_state.click_y if st.session_state.click_y is not None else float(df['y'].mean())

    qx     = c1.number_input("X", value=default_x, format="%.6f", key="qx_input")
    qy     = c2.number_input("Y", value=default_y, format="%.6f", key="qy_input")
    method = c3.selectbox("Method", ["IDW interpolation","Nearest-neighbour","Linear griddata"])
    k      = c4.slider("k neighbours", 3, 20, 6, disabled=(method != "IDW interpolation"))

    if clicked_x is not None:
        st.caption(f"📍 Map click captured: X={clicked_x:.4f}, Y={clicked_y:.4f}")

    if st.button("▶ Query selected point", type="primary"):
        dist, idx = tree.query([qx, qy])
        snap_z = float(df.iloc[idx]['z'])

        if method == "IDW interpolation":
            result_z = idw_single(df['z'].values, tree, qx, qy, k=k)
        elif method == "Nearest-neighbour":
            result_z = snap_z
        else:
            result_z = float(griddata((df['x'], df['y']), df['z'], [(qx, qy)], method='linear')[0])

        r1,r2,r3,r4 = st.columns(4)
        r1.metric("Result Z",           f"{result_z:.6f}")
        r2.metric("Method",             method.split()[0])
        r3.metric("Dist to nearest",    f"{dist:.4f}")
        r4.metric("Snap Z (nearest pt)",f"{snap_z:.6f}")

        st.session_state.query_history.append({"x":qx,"y":qy,"z":result_z,"method":method})

    if st.session_state.query_history:
        st.markdown("#### Query history this session")
        hist_df = pd.DataFrame(st.session_state.query_history)
        st.dataframe(hist_df, use_container_width=True)
        if st.button("Clear history"):
            st.session_state.query_history = []
            st.rerun()


# ─────────────────────────────────────────────────────────────
# TASK 3 — Batch CSV Query  (vectorized IDW)
# ─────────────────────────────────────────────────────────────
def task_batch():
    df   = st.session_state.df
    tree = st.session_state.tree
    st.markdown('<div class="section-header">📂 Batch CSV Query</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Batch CSV Query</div>', unsafe_allow_html=True)
    st.info("Upload a CSV with columns **x** and **y**. IDW is vectorized — handles 100k rows in seconds.")

    method = st.radio("Method", ["IDW interpolation","Nearest-neighbour"], horizontal=True)
    k = st.slider("k (IDW)", 3, 20, 6, disabled=(method != "IDW interpolation"))

    batch_file = st.file_uploader("Upload query CSV", type="csv")
    if not batch_file:
        return

    bdf = pd.read_csv(batch_file)
    if 'x' not in bdf.columns or 'y' not in bdf.columns:
        st.error("CSV must contain columns `x` and `y`."); return

    st.write(f"Preview ({len(bdf):,} rows):", bdf.head())

    if st.button("▶ Run batch query", type="primary"):
        with st.spinner(f"Querying {len(bdf):,} points (vectorized)…"):
            qxy = bdf[['x','y']].values.astype(float)
            if method == "IDW interpolation":
                bdf['z'] = idw_batch(df, tree, qxy, k=k)
            else:
                dists, idxs = tree.query(qxy)
                bdf['z'] = df['z'].values[idxs]
                bdf['distance_to_nearest'] = dists

        st.success(f"Done — {len(bdf):,} points queried.")
        st.dataframe(bdf, use_container_width=True)

        fig = px.scatter(bdf, x='x', y='y', color='z',
                         color_continuous_scale='Plasma', height=400,
                         title="Batch query result map")
        fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        st.plotly_chart(fig, use_container_width=True)

        st.download_button("⬇️ Download results CSV",
                           bdf.to_csv(index=False).encode(),
                           "batch_results.csv", "text/csv")


# ─────────────────────────────────────────────────────────────
# TASK 4 — Radius / Zone Search
# ─────────────────────────────────────────────────────────────
def task_radius():
    df   = st.session_state.df
    tree = st.session_state.tree
    st.markdown('<div class="section-header">🔵 Radius / Zone Search</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Radius / Zone Search</div>', unsafe_allow_html=True)

    c1,c2,c3 = st.columns(3)
    rx = c1.number_input("Centre X", value=float(df['x'].mean()), format="%.6f")
    ry = c2.number_input("Centre Y", value=float(df['y'].mean()), format="%.6f")
    default_r = float((df['x'].max()-df['x'].min())*0.05)
    radius = c3.number_input("Radius", value=default_r, min_value=0.0, format="%.4f")

    if st.button("▶ Search zone", type="primary"):
        idxs = tree.query_ball_point([rx, ry], r=radius)
        if not idxs:
            st.warning("No points in that radius. Try increasing it."); return

        rdf = df.iloc[idxs].copy()
        rdf['distance'] = np.sqrt((rdf['x']-rx)**2 + (rdf['y']-ry)**2)
        rdf = rdf.sort_values('distance').reset_index(drop=True)

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Points found", f"{len(rdf):,}")
        m2.metric("Z mean", f"{rdf['z'].mean():.4f}")
        m3.metric("Z min",  f"{rdf['z'].min():.4f}")
        m4.metric("Z max",  f"{rdf['z'].max():.4f}")

        col_map, col_hist = st.columns(2)
        with col_map:
            fig = px.scatter(rdf, x='x', y='y', color='z',
                             color_continuous_scale='Turbo', height=380,
                             title=f"Zone — radius {radius}")
            fig.add_scatter(x=[rx], y=[ry], mode='markers',
                            marker=dict(color='#f43f5e', size=14, symbol='x-thin', line=dict(width=3)),
                            name='Centre')
            theta = np.linspace(0, 2*np.pi, 120)
            fig.add_scatter(x=rx+radius*np.cos(theta), y=ry+radius*np.sin(theta),
                            mode='lines', line=dict(color='#f43f5e', dash='dash'), name='Radius')
            fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                              font=dict(color='#94a3b8'), margin=dict(l=0,r=0,t=40,b=0))
            st.plotly_chart(fig, use_container_width=True)

        with col_hist:
            fig2 = px.histogram(rdf, x='z', nbins=40, height=380,
                                color_discrete_sequence=['#38bdf8'],
                                title="Z distribution inside zone")
            fig2.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                               font=dict(color='#94a3b8'))
            st.plotly_chart(fig2, use_container_width=True)

        st.dataframe(rdf.head(300), use_container_width=True)
        st.download_button("⬇️ Download zone CSV", rdf.to_csv(index=False).encode(),
                           "zone_results.csv", "text/csv")


# ─────────────────────────────────────────────────────────────
# TASK 5 — Transect Profile
# ─────────────────────────────────────────────────────────────
def task_transect():
    df = st.session_state.df
    tree = st.session_state.tree
    st.markdown('<div class="section-header">📏 Transect Profile</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Transect Profile</div>', unsafe_allow_html=True)
    st.info("Define start and end points. Z values are sampled along the line using vectorized IDW.")

    c1,c2 = st.columns(2)
    with c1:
        x1 = st.number_input("Start X", value=float(df['x'].min()), format="%.6f")
        y1 = st.number_input("Start Y", value=float(df['y'].mean()), format="%.6f")
    with c2:
        x2 = st.number_input("End X", value=float(df['x'].max()), format="%.6f")
        y2 = st.number_input("End Y", value=float(df['y'].mean()), format="%.6f")

    n_samples = st.slider("Sample points along transect", 50, 500, 150)

    if st.button("▶ Generate profile", type="primary"):
        xs = np.linspace(x1, x2, n_samples)
        ys = np.linspace(y1, y2, n_samples)
        qxy = np.column_stack([xs, ys])
        zs = idw_batch(df, tree, qxy, k=6)
        dist_along = np.sqrt((xs-x1)**2 + (ys-y1)**2)

        fig = make_subplots(rows=1, cols=2, subplot_titles=("Cross-section Profile","Transect on Map"))
        fig.add_trace(go.Scatter(
            x=dist_along, y=zs, mode='lines+markers',
            line=dict(color='#38bdf8', width=2),
            marker=dict(size=4),
            fill='tozeroy', fillcolor='rgba(56,189,248,.15)',
            name='Z profile'), row=1, col=1)

        sample = df.sample(min(6000, len(df)), random_state=0)
        fig.add_trace(go.Scatter(
            x=sample['x'], y=sample['y'], mode='markers',
            marker=dict(color=sample['z'], colorscale='Viridis', size=3, opacity=.4),
            name='Data', showlegend=False), row=1, col=2)
        fig.add_trace(go.Scatter(
            x=[x1,x2], y=[y1,y2], mode='lines+markers',
            line=dict(color='#f43f5e', width=3),
            marker=dict(size=10), name='Transect'), row=1, col=2)

        fig.update_layout(height=440, paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        fig.update_xaxes(gridcolor='#1e293b'); fig.update_yaxes(gridcolor='#1e293b')
        st.plotly_chart(fig, use_container_width=True)

        tdf = pd.DataFrame({'distance': dist_along, 'x': xs, 'y': ys, 'z': zs})
        m1,m2,m3 = st.columns(3)
        m1.metric("Transect length", f"{dist_along[-1]:.3f}")
        m2.metric("Z range along line", f"{max(zs)-min(zs):.4f}")
        m3.metric("Mean Z", f"{np.mean(zs):.4f}")
        st.download_button("⬇️ Download profile CSV", tdf.to_csv(index=False).encode(),
                           "transect.csv", "text/csv")


# ─────────────────────────────────────────────────────────────
# TASK 6 — Contour Band Analysis
# ─────────────────────────────────────────────────────────────
def task_band_analysis():
    df = st.session_state.df
    st.markdown('<div class="section-header">🎨 Contour Band Analysis</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Contour Band Analysis</div>', unsafe_allow_html=True)

    z_levels = sorted(df['z'].unique())
    c1,c2 = st.columns(2)
    z_min_sel = c1.selectbox("Z lower bound (inclusive)", z_levels, index=0)
    z_max_sel = c2.selectbox("Z upper bound (inclusive)", z_levels, index=len(z_levels)-1)

    filtered = df[(df['z'] >= z_min_sel) & (df['z'] <= z_max_sel)]
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Points in band", f"{len(filtered):,}")
    m2.metric("% of total",     f"{100*len(filtered)/len(df):.1f}%")
    m3.metric("Z mean",         f"{filtered['z'].mean():.4f}" if len(filtered) else "—")
    m4.metric("Contour levels", f"{filtered['z'].nunique()}")

    band_summary = (
        filtered.groupby('z')
        .agg(count=('z','count'), x_mean=('x','mean'), y_mean=('y','mean'))
        .reset_index()
    )

    col_bar, col_map = st.columns(2)
    with col_bar:
        fig = px.bar(band_summary, x='z', y='count',
                     color='count', color_continuous_scale='Viridis',
                     title="Points per contour level", height=360)
        fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        st.plotly_chart(fig, use_container_width=True)
    with col_map:
        fig2 = px.scatter(filtered, x='x', y='y', color='z',
                          color_continuous_scale='Plasma', opacity=.6,
                          title="Band footprint on map", height=360)
        fig2.update_traces(marker=dict(size=3))
        fig2.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                           font=dict(color='#94a3b8'))
        st.plotly_chart(fig2, use_container_width=True)

    st.dataframe(band_summary, use_container_width=True)
    st.download_button("⬇️ Download band CSV", filtered.to_csv(index=False).encode(),
                       "band_filtered.csv", "text/csv")


# ─────────────────────────────────────────────────────────────
# TASK 7 — Hotspot Detection
# ─────────────────────────────────────────────────────────────
def task_hotspot():
    df = st.session_state.df
    st.markdown('<div class="section-header">🔥 Hotspot & Anomaly Detection</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Hotspot Detection</div>', unsafe_allow_html=True)

    c1,c2 = st.columns(2)
    threshold_pct = c1.slider("Anomaly threshold (top/bottom %)", 1, 20, 5)
    n_extremes    = c2.slider("Show N extreme points", 10, 200, 30)

    z = df['z']
    hi_thresh = np.percentile(z, 100 - threshold_pct)
    lo_thresh = np.percentile(z, threshold_pct)

    highs = df[z >= hi_thresh].copy(); highs['type'] = 'High'
    lows  = df[z <= lo_thresh].copy(); lows['type']  = 'Low'
    extremes = pd.concat([highs, lows])

    m1,m2,m3,m4 = st.columns(4)
    m1.metric("High anomalies", f"{len(highs):,}")
    m2.metric("Low anomalies",  f"{len(lows):,}")
    m3.metric("High threshold", f"{hi_thresh:.4f}")
    m4.metric("Low threshold",  f"{lo_thresh:.4f}")

    sample = df.sample(min(8000, len(df)), random_state=2)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=sample['x'], y=sample['y'], mode='markers',
                             marker=dict(color=sample['z'], colorscale='Greys', size=3, opacity=.3),
                             name='Background'))
    fig.add_trace(go.Scatter(x=highs['x'], y=highs['y'], mode='markers',
                             marker=dict(color='#f43f5e', size=7, symbol='circle'),
                             name=f'High (top {threshold_pct}%)'))
    fig.add_trace(go.Scatter(x=lows['x'], y=lows['y'], mode='markers',
                             marker=dict(color='#38bdf8', size=7, symbol='triangle-down'),
                             name=f'Low (bot {threshold_pct}%)'))
    fig.update_layout(title="Anomaly Map", height=480,
                      paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                      font=dict(color='#94a3b8'))
    st.plotly_chart(fig, use_container_width=True)

    col_h, col_l = st.columns(2)
    with col_h:
        st.markdown("**Top high-Z points**")
        st.dataframe(highs.nlargest(n_extremes, 'z')[['x','y','z']], use_container_width=True)
    with col_l:
        st.markdown("**Top low-Z points**")
        st.dataframe(lows.nsmallest(n_extremes, 'z')[['x','y','z']], use_container_width=True)

    st.download_button("⬇️ Download anomalies CSV", extremes.to_csv(index=False).encode(),
                       "anomalies.csv", "text/csv")


# ─────────────────────────────────────────────────────────────
# TASK 8 — File Comparison
# ─────────────────────────────────────────────────────────────
def task_comparison():
    df  = st.session_state.df
    df2 = st.session_state.df2
    st.markdown('<div class="section-header">⚖️ File Comparison</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → File Comparison</div>', unsafe_allow_html=True)

    if df2 is None:
        st.warning("Please upload a second BNA file on the upload screen.")
        return

    s1, s2 = stat_summary(df), stat_summary(df2)
    src1 = df['source'].iloc[0]; src2 = df2['source'].iloc[0]

    comp = pd.DataFrame({"Statistic": list(s1.keys()),
                         src1: list(s1.values()),
                         src2: list(s2.values())})
    comp['Δ'] = comp[src2] - comp[src1]
    st.dataframe(comp.set_index("Statistic").style.format("{:.4f}"), use_container_width=True)

    col_a, col_b = st.columns(2)
    for col, dff, src, color in [
        (col_a, df, src1, '#38bdf8'), (col_b, df2, src2, '#f43f5e')
    ]:
        with col:
            fig = px.histogram(dff, x='z', nbins=60, title=f"Z — {src}",
                               color_discrete_sequence=[color], height=320)
            fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                              font=dict(color='#94a3b8'))
            st.plotly_chart(fig, use_container_width=True)

    merged = pd.concat([df[['z']].assign(file=src1), df2[['z']].assign(file=src2)])
    fig_v = px.violin(merged, x='file', y='z', box=True, color='file',
                      color_discrete_map={src1:'#38bdf8', src2:'#f43f5e'},
                      title="Z distribution comparison (violin)", height=360)
    fig_v.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                        font=dict(color='#94a3b8'))
    st.plotly_chart(fig_v, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# TASK 9 — 3-D Surface
# ─────────────────────────────────────────────────────────────
def task_3d():
    df = st.session_state.df
    st.markdown('<div class="section-header">🧊 3-D Surface View</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → 3-D Surface</div>', unsafe_allow_html=True)

    c1,c2 = st.columns(2)
    cscale = c1.selectbox("Colour scale", COLORSCALES, index=4)
    res    = c2.slider("Grid resolution", 60, 300, 120, 20)

    sample = df if len(df) <= 60_000 else df.sample(60_000, random_state=0)
    xi = np.linspace(sample['x'].min(), sample['x'].max(), res)
    yi = np.linspace(sample['y'].min(), sample['y'].max(), res)
    XI, YI = np.meshgrid(xi, yi)
    ZI = griddata((sample['x'], sample['y']), sample['z'], (XI, YI), method='linear')

    fig = go.Figure(go.Surface(
        z=ZI, x=XI, y=YI, colorscale=cscale,
        colorbar=dict(title="Z", thickness=14),
        lighting=dict(ambient=.6, diffuse=.8, specular=.4, roughness=.5),
        lightposition=dict(x=100, y=200, z=0),
    ))
    fig.update_layout(
        scene=dict(
            xaxis=dict(backgroundcolor='#0f172a', gridcolor='#1e293b', title='X'),
            yaxis=dict(backgroundcolor='#0f172a', gridcolor='#1e293b', title='Y'),
            zaxis=dict(backgroundcolor='#0f172a', gridcolor='#1e293b', title='Z'),
            bgcolor='#0f172a',
        ),
        paper_bgcolor='#0f172a', font=dict(color='#94a3b8'),
        height=600, margin=dict(l=0,r=0,t=20,b=0),
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption("💡 Click & drag to rotate · Scroll to zoom · Double-click to reset")


# ─────────────────────────────────────────────────────────────
# TASK 10 — Slope & Gradient Map  (NEW)
# ─────────────────────────────────────────────────────────────
def task_slope():
    df = st.session_state.df
    st.markdown('<div class="section-header">📐 Slope & Gradient Map</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Slope & Gradient</div>', unsafe_allow_html=True)
    st.info("Computes the first-order numerical gradient of the interpolated surface to reveal steep changes, ridges, and valleys.")

    c1, c2, c3 = st.columns(3)
    res    = c1.slider("Grid resolution", 80, 400, 180, 20)
    cscale = c2.selectbox("Colour scale", ["Hot","Viridis","Plasma","RdBu"], index=0)
    view   = c3.selectbox("Display", ["Slope magnitude", "dZ/dX (E-W gradient)", "dZ/dY (N-S gradient)", "Aspect (degrees)"])

    sample = df if len(df) <= 60_000 else df.sample(60_000, random_state=0)
    xi = np.linspace(sample['x'].min(), sample['x'].max(), res)
    yi = np.linspace(sample['y'].min(), sample['y'].max(), res)
    XI, YI = np.meshgrid(xi, yi)
    ZI = griddata((sample['x'], sample['y']), sample['z'], (XI, YI), method='linear')

    # Cell size
    dx = (xi[-1] - xi[0]) / (res - 1)
    dy = (yi[-1] - yi[0]) / (res - 1)

    # Gradient
    dz_dy, dz_dx = np.gradient(ZI, dy, dx)
    slope_mag = np.sqrt(dz_dx**2 + dz_dy**2)
    aspect    = np.degrees(np.arctan2(dz_dy, dz_dx)) % 360

    display_map = {
        "Slope magnitude":        (slope_mag, "Slope |∇Z|"),
        "dZ/dX (E-W gradient)":   (dz_dx,     "dZ/dX"),
        "dZ/dY (N-S gradient)":   (dz_dy,     "dZ/dY"),
        "Aspect (degrees)":       (aspect,     "Aspect °"),
    }
    grid_to_show, z_label = display_map[view]

    col_left, col_right = st.columns(2)

    with col_left:
        fig = go.Figure(go.Heatmap(
            z=grid_to_show, x=xi, y=yi,
            colorscale=cscale,
            colorbar=dict(title=z_label, thickness=14),
        ))
        fig.update_layout(title=view, xaxis_title="X", yaxis_title="Y",
                          height=480, margin=dict(l=20,r=20,t=40,b=20),
                          paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        st.plotly_chart(fig, use_container_width=True)

    with col_right:
        # Original surface for comparison
        fig2 = go.Figure(go.Heatmap(
            z=ZI, x=xi, y=yi,
            colorscale="Viridis",
            colorbar=dict(title="Z", thickness=14),
        ))
        fig2.update_layout(title="Original surface (Z)", xaxis_title="X", yaxis_title="Y",
                           height=480, margin=dict(l=20,r=20,t=40,b=20),
                           paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                           font=dict(color='#94a3b8'))
        st.plotly_chart(fig2, use_container_width=True)

    # Summary metrics
    valid = slope_mag[~np.isnan(slope_mag)]
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Mean slope",   f"{valid.mean():.4f}")
    m2.metric("Max slope",    f"{valid.max():.4f}")
    m3.metric("Median slope", f"{np.median(valid):.4f}")
    m4.metric("Steep cells (>75th pct)", f"{(valid > np.percentile(valid,75)).sum():,}")

    # Slope histogram
    fig3 = px.histogram(x=valid.flatten(), nbins=60,
                        color_discrete_sequence=['#f59e0b'],
                        labels={'x':'Slope magnitude'}, title="Slope distribution", height=280)
    fig3.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                       font=dict(color='#94a3b8'))
    st.plotly_chart(fig3, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# TASK 11 — Temporal Animation  (NEW)
# ─────────────────────────────────────────────────────────────
def task_temporal():
    df = st.session_state.df
    st.markdown('<div class="section-header">🎞️ Temporal Animation</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Temporal Animation</div>', unsafe_allow_html=True)
    st.info("Treats each unique Z level as a time step and animates the spatial distribution of points. "
            "Useful when Z encodes time, depth, elevation band, or any ordered dimension.")

    z_levels = sorted(df['z'].unique())
    if len(z_levels) > 200:
        st.warning(f"Found {len(z_levels)} unique Z levels — capping at 60 for animation performance.")
        # Pick evenly spaced subset
        idxs = np.round(np.linspace(0, len(z_levels)-1, 60)).astype(int)
        z_levels = [z_levels[i] for i in idxs]

    c1, c2, c3 = st.columns(3)
    anim_type  = c1.selectbox("Animation type", ["Point positions", "Heatmap frames"])
    cscale     = c2.selectbox("Colour scale", COLORSCALES, index=1)
    frame_dur  = c3.slider("Frame duration (ms)", 100, 1000, 300, 50)

    if st.button("▶ Build animation", type="primary"):
        if anim_type == "Point positions":
            # Scatter animation — each frame = one Z level, cumulative
            frames = []
            for zv in z_levels:
                sub = df[df['z'] == zv]
                frames.append(go.Frame(
                    data=[go.Scatter(
                        x=sub['x'], y=sub['y'], mode='markers',
                        marker=dict(color=sub['z'], colorscale=cscale,
                                    cmin=df['z'].min(), cmax=df['z'].max(),
                                    size=4, opacity=.7),
                    )],
                    name=str(zv),
                    layout=go.Layout(title_text=f"Z level: {zv}")
                ))

            init = df[df['z'] == z_levels[0]]
            fig = go.Figure(
                data=[go.Scatter(
                    x=init['x'], y=init['y'], mode='markers',
                    marker=dict(color=init['z'], colorscale=cscale,
                                cmin=df['z'].min(), cmax=df['z'].max(),
                                size=4, opacity=.7,
                                colorbar=dict(title="Z")),
                )],
                frames=frames,
            )
            fig.update_layout(
                xaxis=dict(range=[df['x'].min(), df['x'].max()]),
                yaxis=dict(range=[df['y'].min(), df['y'].max()]),
                updatemenus=[dict(type="buttons", showactive=False,
                    buttons=[
                        dict(label="▶ Play",  method="animate",
                             args=[None, {"frame":{"duration":frame_dur,"redraw":True},
                                          "fromcurrent":True}]),
                        dict(label="⏸ Pause", method="animate",
                             args=[[None], {"frame":{"duration":0},"mode":"immediate"}])
                    ])],
                sliders=[dict(
                    steps=[dict(method="animate", args=[[str(z)],{"mode":"immediate","frame":{"duration":frame_dur}}],
                                label=str(z)) for z in z_levels],
                    transition=dict(duration=0), x=0.05, y=0, len=0.9,
                    currentvalue=dict(prefix="Z: ", visible=True),
                )],
                height=560, paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                font=dict(color='#94a3b8'),
            )

        else:
            # Heatmap animation — gridded frames
            res = 120
            xi = np.linspace(df['x'].min(), df['x'].max(), res)
            yi = np.linspace(df['y'].min(), df['y'].max(), res)
            XI, YI = np.meshgrid(xi, yi)

            frames = []
            with st.spinner("Interpolating frames…"):
                for zv in z_levels:
                    sub = df[df['z'] <= zv]          # cumulative up to this level
                    if len(sub) < 4:
                        continue
                    ZI = griddata((sub['x'], sub['y']), sub['z'], (XI, YI), method='linear')
                    frames.append(go.Frame(
                        data=[go.Heatmap(z=ZI, x=xi, y=yi, colorscale=cscale,
                                         zmin=df['z'].min(), zmax=df['z'].max())],
                        name=str(zv),
                        layout=go.Layout(title_text=f"Cumulative up to Z={zv}")
                    ))

            first_sub = df[df['z'] <= z_levels[0]]
            ZI0 = griddata((first_sub['x'], first_sub['y']), first_sub['z'],
                           (XI, YI), method='linear') if len(first_sub) >= 4 else np.full((res,res), np.nan)

            fig = go.Figure(
                data=[go.Heatmap(z=ZI0, x=xi, y=yi, colorscale=cscale,
                                 zmin=df['z'].min(), zmax=df['z'].max(),
                                 colorbar=dict(title="Z"))],
                frames=frames,
            )
            fig.update_layout(
                updatemenus=[dict(type="buttons", showactive=False,
                    buttons=[
                        dict(label="▶ Play",  method="animate",
                             args=[None, {"frame":{"duration":frame_dur,"redraw":True},"fromcurrent":True}]),
                        dict(label="⏸ Pause", method="animate",
                             args=[[None], {"frame":{"duration":0},"mode":"immediate"}])
                    ])],
                sliders=[dict(
                    steps=[dict(method="animate",
                                args=[[str(z)],{"mode":"immediate","frame":{"duration":frame_dur}}],
                                label=str(z)) for z in z_levels],
                    transition=dict(duration=0), x=0.05, y=0, len=0.9,
                    currentvalue=dict(prefix="Z≤ ", visible=True),
                )],
                height=560, paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                font=dict(color='#94a3b8'),
            )

        st.plotly_chart(fig, use_container_width=True)
        st.caption(f"🎞️ {len(z_levels)} frames  ·  {frame_dur} ms/frame  ·  Use the Play button or drag the slider")


# ─────────────────────────────────────────────────────────────
# TASK 12 — Data Quality & Outlier Detection  (NEW)
# ─────────────────────────────────────────────────────────────
def task_quality():
    df   = st.session_state.df
    tree = st.session_state.tree
    st.markdown('<div class="section-header">🔍 Data Quality Report</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Data Quality</div>', unsafe_allow_html=True)

    # ── 1. Duplicate detection ───────────────────────────────
    dupes = df.duplicated(subset=['x','y'], keep=False)
    n_dupes = dupes.sum()

    # ── 2. IQR-based Z outliers ──────────────────────────────
    Q1, Q3 = df['z'].quantile(0.25), df['z'].quantile(0.75)
    IQR = Q3 - Q1
    lo_fence = Q1 - 3 * IQR
    hi_fence = Q3 + 3 * IQR
    z_outliers = df[(df['z'] < lo_fence) | (df['z'] > hi_fence)]

    # ── 3. Spatial gaps — find points with unusually large nn distance ──
    dists, _ = tree.query(df[['x','y']].values, k=2)
    nn_dists = dists[:, 1]
    gap_thresh = np.percentile(nn_dists, 95)
    sparse_pts = df[nn_dists > gap_thresh].copy()
    sparse_pts['nn_dist'] = nn_dists[nn_dists > gap_thresh]

    # ── 4. Z distribution normality test ────────────────────
    sample_z = df['z'].sample(min(5000, len(df)), random_state=0)
    _, p_value = scipy_stats.shapiro(sample_z)
    is_normal = p_value > 0.05

    # ── Summary badges ───────────────────────────────────────
    st.markdown("### Quality Summary")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        if n_dupes == 0:
            st.markdown('<span class="badge-ok">✓ No duplicates</span>', unsafe_allow_html=True)
        else:
            st.markdown(f'<span class="badge-warn">⚠ {n_dupes} duplicate XY pairs</span>', unsafe_allow_html=True)
        st.metric("Duplicate XY pairs", f"{n_dupes:,}")

    with col2:
        if len(z_outliers) == 0:
            st.markdown('<span class="badge-ok">✓ No Z outliers</span>', unsafe_allow_html=True)
        else:
            st.markdown(f'<span class="badge-warn">⚠ {len(z_outliers)} Z outliers (3×IQR)</span>', unsafe_allow_html=True)
        st.metric("Z outliers (3×IQR)", f"{len(z_outliers):,}")

    with col3:
        st.markdown(f'<span class="badge-warn">⚠ {len(sparse_pts)} sparse regions</span>'
                    if len(sparse_pts) > 0 else '<span class="badge-ok">✓ Even coverage</span>',
                    unsafe_allow_html=True)
        st.metric("Sparse regions (top-5% gap)", f"{len(sparse_pts):,}")

    with col4:
        label = "Normal-ish" if is_normal else "Non-normal"
        badge = "badge-ok" if is_normal else "badge-warn"
        st.markdown(f'<span class="{badge}">{"✓" if is_normal else "⚠"} {label} (Shapiro p={p_value:.3f})</span>',
                    unsafe_allow_html=True)
        st.metric("Shapiro-Wilk p-value", f"{p_value:.4f}")

    st.markdown("---")

    # ── Plots ────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs(["Z Outliers", "Spatial Gaps", "NN Distance Distribution", "Z Box Plot"])

    with tab1:
        if len(z_outliers) == 0:
            st.success("No Z outliers detected at 3×IQR threshold.")
        else:
            sample_bg = df.sample(min(5000, len(df)), random_state=3)
            fig = go.Figure()
            fig.add_scatter(x=sample_bg['x'], y=sample_bg['y'], mode='markers',
                            marker=dict(color='#334155', size=3), name='Normal points')
            fig.add_scatter(x=z_outliers['x'], y=z_outliers['y'], mode='markers',
                            marker=dict(color='#f43f5e', size=8, symbol='x',
                                        line=dict(width=2)), name='Z outliers')
            fig.update_layout(title=f"Z outlier locations (fence: [{lo_fence:.3f}, {hi_fence:.3f}])",
                              height=420, paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                              font=dict(color='#94a3b8'))
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(z_outliers[['x','y','z']].sort_values('z'), use_container_width=True)
            st.download_button("⬇️ Download Z outliers CSV",
                               z_outliers.to_csv(index=False).encode(),
                               "z_outliers.csv", "text/csv")

    with tab2:
        if len(sparse_pts) == 0:
            st.success("No significant spatial gaps detected.")
        else:
            fig = go.Figure()
            fig.add_scatter(x=df['x'], y=df['y'], mode='markers',
                            marker=dict(color='#1e293b', size=2), name='All points')
            fig.add_scatter(x=sparse_pts['x'], y=sparse_pts['y'], mode='markers',
                            marker=dict(color='#f59e0b', size=8, symbol='diamond'),
                            name='Sparse / gap points')
            fig.update_layout(title="Points in sparse regions (large NN distance)",
                              height=420, paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                              font=dict(color='#94a3b8'))
            st.plotly_chart(fig, use_container_width=True)
            st.download_button("⬇️ Download sparse points CSV",
                               sparse_pts.to_csv(index=False).encode(),
                               "sparse_points.csv", "text/csv")

    with tab3:
        fig = px.histogram(x=nn_dists, nbins=80,
                           labels={'x':'Nearest-neighbour distance'},
                           color_discrete_sequence=['#6366f1'],
                           title="Nearest-neighbour distance distribution", height=360)
        fig.add_vline(x=gap_thresh, line_dash="dash", line_color="#f59e0b",
                      annotation_text="95th pct gap threshold", annotation_position="top right")
        fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        st.plotly_chart(fig, use_container_width=True)

    with tab4:
        fig = go.Figure()
        fig.add_trace(go.Box(y=df['z'], name='All Z', marker_color='#38bdf8',
                             boxmean='sd'))
        fig.add_hline(y=lo_fence, line_dash="dash", line_color="#f43f5e",
                      annotation_text="Lower fence (3×IQR)")
        fig.add_hline(y=hi_fence, line_dash="dash", line_color="#f43f5e",
                      annotation_text="Upper fence (3×IQR)")
        fig.update_layout(title="Z box plot with outlier fences", height=420,
                          paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'))
        st.plotly_chart(fig, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# TASK 13 — Export & Report  (now includes Excel)
# ─────────────────────────────────────────────────────────────
def task_export():
    df = st.session_state.df
    st.markdown('<div class="section-header">📤 Export & Report</div>', unsafe_allow_html=True)
    st.markdown('<div class="breadcrumb">Home → Export & Report</div>', unsafe_allow_html=True)

    src = df['source'].iloc[0]
    s   = stat_summary(df)

    st.markdown("### Summary statistics")
    stat_df = pd.DataFrame({"Statistic": list(s.keys()), "Value": list(s.values())})
    st.dataframe(stat_df.set_index("Statistic").style.format("{:.6f}"), use_container_width=True)

    level_tbl = (df.groupby('z')
                   .agg(point_count=('z','count'),
                        x_min=('x','min'), x_max=('x','max'),
                        y_min=('y','min'), y_max=('y','max'))
                   .reset_index()
                   .rename(columns={'z':'contour_value'}))

    st.markdown("### Contour level table")
    st.dataframe(level_tbl, use_container_width=True)

    col_a, col_b = st.columns(2)
    with col_a:
        fig = px.histogram(df, x='z', nbins=60, title="Z distribution",
                           color_discrete_sequence=['#38bdf8'])
        fig.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                          font=dict(color='#94a3b8'), height=300)
        st.plotly_chart(fig, use_container_width=True)
    with col_b:
        fig2 = px.box(df, y='z', title="Z box plot",
                      color_discrete_sequence=['#6366f1'])
        fig2.update_layout(paper_bgcolor='#0f172a', plot_bgcolor='#0f172a',
                           font=dict(color='#94a3b8'), height=300)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("### Downloads")
    dl1, dl2, dl3, dl4 = st.columns(4)

    dl1.download_button("⬇️ Full dataset CSV", df.to_csv(index=False).encode(),
                        "bna_data.csv", "text/csv")
    dl2.download_button("⬇️ Statistics CSV",  stat_df.to_csv(index=False).encode(),
                        "bna_stats.csv", "text/csv")
    dl3.download_button("⬇️ Level table CSV", level_tbl.to_csv(index=False).encode(),
                        "bna_levels.csv", "text/csv")

    # ── Excel export ─────────────────────────────────────────
    if dl4.button("📊 Build Excel report"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.formatting.rule import ColorScaleRule

            wb = openpyxl.Workbook()

            # ── Sheet 1: Summary ──────────────────────────────
            ws1 = wb.active
            ws1.title = "Summary"
            hdr_fill  = PatternFill("solid", start_color="0F172A")
            hdr_font  = Font(bold=True, color="38BDF8", name="Arial")
            val_font  = Font(name="Arial", color="F1F5F9")
            thin      = Border(
                left=Side(style='thin', color='334155'),
                right=Side(style='thin', color='334155'),
                top=Side(style='thin', color='334155'),
                bottom=Side(style='thin', color='334155'),
            )

            ws1["A1"] = "BNA Contour Explorer — Export Report"
            ws1["A1"].font = Font(bold=True, size=14, color="38BDF8", name="Arial")
            ws1["A2"] = f"Source file: {src}"
            ws1["A2"].font = Font(italic=True, color="94A3B8", name="Arial")
            ws1.append([])

            ws1.append(["Statistic", "Value"])
            for cell in ws1[ws1.max_row]:
                cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin
                cell.alignment = Alignment(horizontal="center")

            for k, v in s.items():
                ws1.append([k, round(v, 6)])
                for cell in ws1[ws1.max_row]:
                    cell.font = val_font; cell.border = thin
                    cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right")

            ws1.column_dimensions["A"].width = 18
            ws1.column_dimensions["B"].width = 18

            # ── Sheet 2: Contour Level Table ──────────────────
            ws2 = wb.create_sheet("Contour Levels")
            headers = list(level_tbl.columns)
            ws2.append(headers)
            for cell in ws2[1]:
                cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin
                cell.alignment = Alignment(horizontal="center")

            for _, row in level_tbl.iterrows():
                ws2.append([round(v, 6) if isinstance(v, float) else v for v in row])
                for cell in ws2[ws2.max_row]:
                    cell.font = val_font; cell.border = thin

            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 16

            # Colour scale on contour_value column (A)
            ws2.conditional_formatting.add(
                f"A2:A{ws2.max_row}",
                ColorScaleRule(start_type='min', start_color='0EA5E9',
                               end_type='max',   end_color='F43F5E')
            )

            # ── Sheet 3: Raw data (first 50k rows) ───────────
            ws3 = wb.create_sheet("Raw Data")
            ws3.append(["x", "y", "z"])
            for cell in ws3[1]:
                cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin

            export_df = df[['x','y','z']].head(50_000)
            for _, row in export_df.iterrows():
                ws3.append([row['x'], row['y'], row['z']])

            for col in ["A","B","C"]:
                ws3.column_dimensions[col].width = 16

            # ── Sheet 4: Formulas / Stats summary ────────────
            ws4 = wb.create_sheet("Formula Check")
            ws4["A1"] = "Stat"
            ws4["B1"] = "Value (formula)"
            ws4["A1"].font = hdr_font; ws4["B1"].font = hdr_font

            # Write raw Z values to a hidden helper column for Excel formulas
            ws4["A2"] = "Count";   ws4["B2"] = f"=COUNTA('Raw Data'!C:C)-1"
            ws4["A3"] = "Min Z";   ws4["B3"] = f"=MIN('Raw Data'!C2:C{len(export_df)+1})"
            ws4["A4"] = "Max Z";   ws4["B4"] = f"=MAX('Raw Data'!C2:C{len(export_df)+1})"
            ws4["A5"] = "Mean Z";  ws4["B5"] = f"=AVERAGE('Raw Data'!C2:C{len(export_df)+1})"
            ws4["A6"] = "StDev Z"; ws4["B6"] = f"=STDEV('Raw Data'!C2:C{len(export_df)+1})"

            for row in ws4.iter_rows(min_row=2, max_row=6):
                for cell in row:
                    cell.font = val_font
            ws4.column_dimensions["A"].width = 14
            ws4.column_dimensions["B"].width = 20

            # ── Save ─────────────────────────────────────────
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            st.download_button(
                "⬇️ Download Excel report (.xlsx)",
                data=buf.getvalue(),
                file_name="bna_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success("Excel report ready — 4 sheets: Summary · Contour Levels · Raw Data · Formula Check")
        except ImportError:
            st.error("openpyxl not installed. Run: pip install openpyxl")

    # Plain-text report
    report = f"""BNA CONTOUR EXPLORER — REPORT
==============================
File:      {src}

STATISTICS
----------
{"".join(f"{k:<14}{v:.6f}\n" for k,v in s.items())}

CONTOUR LEVELS  ({df['z'].nunique()} unique)
--------------
{level_tbl.to_string(index=False)}
"""
    st.download_button("⬇️ Plain-text report", report.encode(),
                       "bna_report.txt", "text/plain")


# ─────────────────────────────────────────────────────────────
# ROUTER
# ─────────────────────────────────────────────────────────────
TASK_MAP = {
    "Interactive Map":      task_map,
    "Point Query":          task_point_query,
    "Batch CSV Query":      task_batch,
    "Radius / Zone Search": task_radius,
    "Transect Profile":     task_transect,
    "Contour Band Analysis":task_band_analysis,
    "Hotspot Detection":    task_hotspot,
    "File Comparison":      task_comparison,
    "3-D Surface":          task_3d,
    "Slope & Gradient":     task_slope,
    "Temporal Animation":   task_temporal,
    "Data Quality":         task_quality,
    "Export & Report":      task_export,
}

st.markdown('<h2 style="color:#f1f5f9;font-weight:800;margin-bottom:0">🗺️ BNA Contour Explorer</h2>', unsafe_allow_html=True)
st.markdown('<hr style="border-color:#1e293b;margin-top:8px">', unsafe_allow_html=True)

if st.session_state.df is None:
    upload_screen()
elif st.session_state.task is None:
    task_hub()
else:
    back_button()
    TASK_MAP[st.session_state.task]()
